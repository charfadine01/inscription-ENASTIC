"""
Fenêtre principale de l'application desktop ENASTIC
"""

from PyQt6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                              QTabWidget, QLabel, QPushButton, QTableWidget,
                              QTableWidgetItem, QLineEdit, QComboBox, QMessageBox,
                              QFileDialog, QStatusBar, QFrame, QHeaderView, QDateEdit,
                              QFormLayout, QDialog, QDialogButtonBox, QTextEdit, QProgressBar)
from PyQt6.QtCore import Qt, QDate, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QIcon, QPixmap
from datetime import datetime
import os


class MainWindow(QMainWindow):
    """Fenêtre principale de l'application"""

    def __init__(self, db_manager, user_data):
        super().__init__()
        self.db_manager = db_manager
        self.user = user_data
        self.init_ui()
        self.load_initial_data()

    def init_ui(self):
        """Initialise l'interface utilisateur"""
        self.setWindowTitle(f"ENASTIC - Gestion des Inscriptions (Connecté: {self.user['nom_complet']})")
        self.setGeometry(100, 100, 1400, 900)

        # Définir l'icône de la fenêtre
        icon_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets', 'logo.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        # Style global
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f7ff;
            }
            QTabWidget::pane {
                border: 1px solid #E0E0E0;
                background: white;
                border-radius: 8px;
            }
            QTabBar::tab {
                background: #E5F0FF;
                color: #4A90E2;
                padding: 12px 24px;
                margin-right: 4px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: 500;
            }
            QTabBar::tab:selected {
                background: white;
                color: #4A90E2;
                font-weight: bold;
            }
            QPushButton {
                background-color: #4A90E2;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #3B82F6;
            }
            QPushButton:pressed {
                background-color: #2563EB;
            }
            QPushButton:disabled {
                background-color: #CCCCCC;
            }
            QLineEdit, QComboBox, QDateEdit {
                padding: 8px;
                border: 2px solid #E0E0E0;
                border-radius: 6px;
                background: white;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {
                border: 2px solid #4A90E2;
            }
            QTableWidget {
                border: 1px solid #E0E0E0;
                border-radius: 6px;
                gridline-color: #F0F0F0;
            }
            QHeaderView::section {
                background-color: #4A90E2;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
            }
            QLabel {
                color: #374151;
            }
        """)

        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        # En-tête
        header = self.create_header()
        main_layout.addWidget(header)

        # Tabs
        self.tabs = QTabWidget()
        self.tabs.addTab(self.create_dashboard_tab(), "📊 Tableau de Bord")
        self.tabs.addTab(self.create_students_tab(), "👥 Étudiants")
        self.tabs.addTab(self.create_certificates_tab(), "🎓 Certificats")
        self.tabs.addTab(self.create_import_tab(), "📥 Import Excel")
        self.tabs.addTab(self.create_settings_tab(), "⚙️ Paramètres")

        main_layout.addWidget(self.tabs)

        central_widget.setLayout(main_layout)

        # Barre de statut
        self.statusBar().showMessage(f"Connecté en tant que {self.user['nom_complet']} - Site: {self.user['site']}")

    def create_header(self):
        """Crée l'en-tête de l'application"""
        header = QFrame()
        header.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 12px;
                padding: 15px;
            }
        """)

        layout = QHBoxLayout()

        # Logo ENASTIC
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets', 'logo.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path)
            scaled_pixmap = pixmap.scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(scaled_pixmap)
            layout.addWidget(logo_label)

        # Titre
        title_layout = QVBoxLayout()
        title = QLabel("ENASTIC")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #4A90E2;")

        subtitle = QLabel("École Nationale Supérieure des Technologies de l'Information et de la Communication")
        subtitle.setStyleSheet("font-size: 12px; color: #6B7280;")

        title_layout.addWidget(title)
        title_layout.addWidget(subtitle)

        # Infos utilisateur
        user_info = QLabel(f"👤 {self.user['nom_complet']}\n🏫 {self.user['site']}")
        user_info.setStyleSheet("font-size: 13px; color: #374151;")
        user_info.setAlignment(Qt.AlignmentFlag.AlignRight)

        # Bouton déconnexion
        logout_btn = QPushButton("🚪 Déconnexion")
        logout_btn.setStyleSheet("""
            QPushButton {
                background-color: #DC2626;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #B91C1C;
            }
        """)
        logout_btn.clicked.connect(self.handle_logout)

        layout.addLayout(title_layout)
        layout.addStretch()
        layout.addWidget(user_info)
        layout.addWidget(logout_btn)

        header.setLayout(layout)
        return header

    def create_dashboard_tab(self):
        """Crée l'onglet tableau de bord"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)

        # Statistiques
        stats_frame = QFrame()
        stats_frame.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 10px;
                padding: 20px;
            }
        """)

        stats_layout = QVBoxLayout()
        stats_title = QLabel("📊 Statistiques")
        stats_title.setStyleSheet("font-size: 18px; font-weight: bold; color: #4A90E2;")
        stats_layout.addWidget(stats_title)

        # Grille de statistiques
        self.stats_grid = QHBoxLayout()
        stats_layout.addLayout(self.stats_grid)

        stats_frame.setLayout(stats_layout)
        layout.addWidget(stats_frame)

        # Activités récentes
        activities_frame = QFrame()
        activities_frame.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 10px;
                padding: 20px;
            }
        """)

        activities_layout = QVBoxLayout()
        activities_title = QLabel("📝 Activités Récentes")
        activities_title.setStyleSheet("font-size: 18px; font-weight: bold; color: #4A90E2;")
        activities_layout.addWidget(activities_title)

        self.activities_list = QTableWidget()
        self.activities_list.setColumnCount(4)
        self.activities_list.setHorizontalHeaderLabels(["Date", "Action", "Détails", "Utilisateur"])
        self.activities_list.horizontalHeader().setStretchLastSection(True)
        activities_layout.addWidget(self.activities_list)

        activities_frame.setLayout(activities_layout)
        layout.addWidget(activities_frame)

        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def create_students_tab(self):
        """Crée l'onglet gestion des étudiants"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)

        # Barre de recherche et filtres
        search_layout = QHBoxLayout()

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Rechercher par nom, prénom ou matricule...")
        self.search_input.textChanged.connect(self.filter_students)

        self.site_filter = QComboBox()
        self.site_filter.addItem("Tous les sites")
        self.site_filter.addItems(["Ndjamena", "Sarh", "Amdjarass"])
        self.site_filter.currentTextChanged.connect(self.filter_students)

        self.status_filter = QComboBox()
        self.status_filter.addItem("Tous les statuts")
        self.status_filter.addItems(["En attente", "Validé", "Rejeté"])
        self.status_filter.currentTextChanged.connect(self.filter_students)

        search_layout.addWidget(self.search_input, 3)
        search_layout.addWidget(self.site_filter, 1)
        search_layout.addWidget(self.status_filter, 1)

        layout.addLayout(search_layout)

        # Boutons d'action
        buttons_layout = QHBoxLayout()

        add_btn = QPushButton("➕ Nouvel Étudiant")
        add_btn.clicked.connect(self.add_student)

        refresh_btn = QPushButton("🔄 Actualiser")
        refresh_btn.clicked.connect(self.load_students)

        export_btn = QPushButton("📤 Exporter Excel")
        export_btn.clicked.connect(self.export_students)

        buttons_layout.addWidget(add_btn)
        buttons_layout.addWidget(refresh_btn)
        buttons_layout.addWidget(export_btn)
        buttons_layout.addStretch()

        layout.addLayout(buttons_layout)

        # Table des étudiants
        self.students_table = QTableWidget()
        self.students_table.setColumnCount(10)
        self.students_table.setHorizontalHeaderLabels([
            "ID", "Matricule", "Nom", "Prénoms", "Genre", "Classe",
            "Site", "Année", "Statut", "Actions"
        ])
        self.students_table.hideColumn(0)  # Cacher l'ID
        self.students_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.students_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.students_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

        layout.addWidget(self.students_table)

        widget.setLayout(layout)
        return widget

    def create_certificates_tab(self):
        """Crée l'onglet génération de certificats"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)

        # Section génération
        gen_frame = QFrame()
        gen_frame.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 10px;
                padding: 20px;
            }
        """)

        gen_layout = QVBoxLayout()
        title = QLabel("🎓 Génération de Certificats")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #4A90E2; margin-bottom: 15px;")
        gen_layout.addWidget(title)

        # Formulaire
        form_layout = QFormLayout()

        self.cert_site = QComboBox()
        self.cert_site.addItems(["Ndjamena", "Sarh", "Amdjarass"])
        self.cert_site.currentTextChanged.connect(self.load_classes_for_cert)

        self.cert_classe = QComboBox()

        self.cert_annee = QLineEdit(self.db_manager.get_config('annee_academique_actuelle') or '2025-2026')

        form_layout.addRow("Site:", self.cert_site)
        form_layout.addRow("Classe:", self.cert_classe)
        form_layout.addRow("Année:", self.cert_annee)

        gen_layout.addLayout(form_layout)

        # Boutons de génération
        cert_buttons = QHBoxLayout()

        generate_btn = QPushButton("📄 Générer les Certificats")
        generate_btn.clicked.connect(self.generate_certificates)

        preview_btn = QPushButton("👁️ Prévisualiser")
        preview_btn.clicked.connect(self.preview_certificates)

        cert_buttons.addWidget(preview_btn)
        cert_buttons.addWidget(generate_btn)
        cert_buttons.addStretch()

        gen_layout.addLayout(cert_buttons)

        # Barre de progression
        self.cert_progress = QProgressBar()
        self.cert_progress.hide()
        gen_layout.addWidget(self.cert_progress)

        gen_frame.setLayout(gen_layout)
        layout.addWidget(gen_frame)

        # Historique
        history_frame = QFrame()
        history_frame.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 10px;
                padding: 20px;
            }
        """)

        history_layout = QVBoxLayout()
        history_title = QLabel("📋 Historique des Certificats Générés")
        history_title.setStyleSheet("font-size: 16px; font-weight: bold; color: #4A90E2;")
        history_layout.addWidget(history_title)

        self.cert_history_table = QTableWidget()
        self.cert_history_table.setColumnCount(5)
        self.cert_history_table.setHorizontalHeaderLabels(["Date", "Étudiant", "Type", "Fichier", "Actions"])
        history_layout.addWidget(self.cert_history_table)

        history_frame.setLayout(history_layout)
        layout.addWidget(history_frame)

        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def create_import_tab(self):
        """Crée l'onglet import Excel"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)

        # Instructions
        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame {
                background: #E3F2FD;
                border-left: 4px solid #2196F3;
                border-radius: 8px;
                padding: 15px;
            }
        """)

        info_layout = QVBoxLayout()
        info_title = QLabel("ℹ️ Comment importer une liste d'étudiants")
        info_title.setStyleSheet("font-weight: bold; font-size: 14px; color: #1976D2;")

        info_text = QLabel(
            "1. Téléchargez le template Excel ci-dessous\n"
            "2. Remplissez-le avec les données des étudiants\n"
            "3. Uploadez le fichier et générez les certificats automatiquement"
        )
        info_text.setStyleSheet("color: #1565C0; line-height: 1.5;")

        info_layout.addWidget(info_title)
        info_layout.addWidget(info_text)
        info_frame.setLayout(info_layout)

        layout.addWidget(info_frame)

        # Télécharger template
        template_btn = QPushButton("📥 Télécharger le Template Excel")
        template_btn.clicked.connect(self.download_template)
        layout.addWidget(template_btn)

        # Sélection fichier
        file_layout = QHBoxLayout()

        self.import_file_label = QLabel("Aucun fichier sélectionné")
        self.import_file_label.setStyleSheet("padding: 10px; background: white; border: 2px dashed #CCCCCC; border-radius: 6px;")

        select_file_btn = QPushButton("📁 Sélectionner un fichier Excel")
        select_file_btn.clicked.connect(self.select_import_file)

        file_layout.addWidget(self.import_file_label, 3)
        file_layout.addWidget(select_file_btn, 1)

        layout.addLayout(file_layout)

        # Options
        self.import_compile_classe = QComboBox()
        self.import_compile_classe.addItems(["Compiler par classe", "Ne pas compiler"])

        layout.addWidget(self.import_compile_classe)

        # Bouton d'import
        import_btn = QPushButton("🚀 Importer et Générer les Certificats")
        import_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                font-size: 16px;
                padding: 15px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        import_btn.clicked.connect(self.import_excel)
        layout.addWidget(import_btn)

        # Résultats
        self.import_results = QTextEdit()
        self.import_results.setReadOnly(True)
        self.import_results.hide()
        layout.addWidget(self.import_results)

        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def create_settings_tab(self):
        """Crée l'onglet paramètres"""
        widget = QWidget()
        layout = QVBoxLayout()

        settings_frame = QFrame()
        settings_frame.setStyleSheet("""
            QFrame {
                background: white;
                border-radius: 10px;
                padding: 20px;
            }
        """)

        settings_layout = QFormLayout()

        # Année académique
        self.settings_annee = QLineEdit(self.db_manager.get_config('annee_academique_actuelle') or '2025-2026')

        # Boutons
        save_btn = QPushButton("💾 Enregistrer les Paramètres")
        save_btn.clicked.connect(self.save_settings)

        backup_btn = QPushButton("💾 Sauvegarder la Base de Données")
        backup_btn.clicked.connect(self.backup_database)

        settings_layout.addRow("Année Académique:", self.settings_annee)
        settings_layout.addRow("", save_btn)
        settings_layout.addRow("", backup_btn)

        settings_frame.setLayout(settings_layout)
        layout.addWidget(settings_frame)
        layout.addStretch()

        widget.setLayout(layout)
        return widget

    # ==================== MÉTHODES D'ACTION ====================

    def load_initial_data(self):
        """Charge les données initiales"""
        self.load_statistics()
        self.load_students()
        self.load_classes_for_cert()

    def load_statistics(self):
        """Charge et affiche les statistiques"""
        stats = self.db_manager.get_statistics(
            site=self.user['site'] if self.user['site'] != 'Tous' else None
        )

        # Effacer les widgets existants
        while self.stats_grid.count():
            child = self.stats_grid.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        # Créer les cartes de stats
        self.add_stat_card("Total Étudiants", str(stats['total']), "#4A90E2")
        self.add_stat_card("En attente", str(stats['par_statut'].get('En attente', 0)), "#FF9800")
        self.add_stat_card("Validés", str(stats['par_statut'].get('Validé', 0)), "#28a745")
        self.add_stat_card("Masculin", str(stats['par_genre'].get('M', 0)), "#2196F3")
        self.add_stat_card("Féminin", str(stats['par_genre'].get('F', 0)), "#E91E63")

    def add_stat_card(self, label, value, color):
        """Ajoute une carte de statistique"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background: {color};
                border-radius: 8px;
                padding: 15px;
            }}
        """)

        layout = QVBoxLayout()

        value_label = QLabel(value)
        value_label.setStyleSheet("color: white; font-size: 32px; font-weight: bold;")

        text_label = QLabel(label)
        text_label.setStyleSheet("color: white; font-size: 14px;")

        layout.addWidget(value_label)
        layout.addWidget(text_label)

        card.setLayout(layout)
        self.stats_grid.addWidget(card)

    def load_students(self):
        """Charge la liste des étudiants"""
        filters = {}

        # Appliquer les filtres
        if self.site_filter.currentText() != "Tous les sites":
            filters['site_formation'] = self.site_filter.currentText()

        if self.status_filter.currentText() != "Tous les statuts":
            filters['statut'] = self.status_filter.currentText()

        if self.search_input.text():
            filters['search'] = self.search_input.text()

        # Si l'utilisateur n'est pas admin, filtrer par son site
        if self.user['site'] != 'Tous':
            filters['site_formation'] = self.user['site']

        students = self.db_manager.get_etudiants(filters)

        # Remplir le tableau
        self.students_table.setRowCount(len(students))

        for row, student in enumerate(students):
            self.students_table.setItem(row, 0, QTableWidgetItem(str(student['id'])))
            self.students_table.setItem(row, 1, QTableWidgetItem(student['matricule']))
            self.students_table.setItem(row, 2, QTableWidgetItem(student['nom']))
            self.students_table.setItem(row, 3, QTableWidgetItem(student['prenoms']))
            self.students_table.setItem(row, 4, QTableWidgetItem(student['genre']))
            self.students_table.setItem(row, 5, QTableWidgetItem(student['classe']))
            self.students_table.setItem(row, 6, QTableWidgetItem(student['site_formation']))
            self.students_table.setItem(row, 7, QTableWidgetItem(student['annee_academique']))

            # Statut coloré
            status_item = QTableWidgetItem(student['statut'])
            if student['statut'] == 'Validé':
                status_item.setBackground(QColor("#d4edda"))
            elif student['statut'] == 'Rejeté':
                status_item.setBackground(QColor("#f8d7da"))
            else:
                status_item.setBackground(QColor("#fff3cd"))
            self.students_table.setItem(row, 8, status_item)

            # Boutons d'action
            actions_widget = QWidget()
            actions_layout = QHBoxLayout()
            actions_layout.setContentsMargins(5, 2, 5, 2)

            edit_btn = QPushButton("✏️")
            edit_btn.setMaximumWidth(40)
            edit_btn.clicked.connect(lambda checked, s_id=student['id']: self.edit_student(s_id))

            delete_btn = QPushButton("🗑️")
            delete_btn.setMaximumWidth(40)
            delete_btn.setStyleSheet("QPushButton { background-color: #DC2626; }")
            delete_btn.clicked.connect(lambda checked, s_id=student['id']: self.delete_student(s_id))

            actions_layout.addWidget(edit_btn)
            actions_layout.addWidget(delete_btn)
            actions_widget.setLayout(actions_layout)

            self.students_table.setCellWidget(row, 9, actions_widget)

    def filter_students(self):
        """Filtre les étudiants"""
        self.load_students()

    def add_student(self):
        """Ouvre le dialogue d'ajout d'étudiant"""
        dialog = AddStudentDialog(self.db_manager, self.user, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Recharger la liste des étudiants et les statistiques
            self.load_students()
            self.load_statistics()

    def edit_student(self, student_id):
        """Édite un étudiant"""
        # Récupérer les données de l'étudiant
        student = self.db_manager.get_etudiant_by_id(student_id)
        if not student:
            QMessageBox.warning(self, "Erreur", "Étudiant introuvable")
            return

        dialog = EditStudentDialog(self.db_manager, self.user, student, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Recharger la liste
            self.load_students()
            self.load_statistics()

    def delete_student(self, student_id):
        """Supprime un étudiant"""
        reply = QMessageBox.question(self, "Confirmation",
                                     "Êtes-vous sûr de vouloir supprimer cet étudiant ?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            if self.db_manager.delete_etudiant(student_id, self.user['id']):
                QMessageBox.information(self, "Succès", "Étudiant supprimé avec succès")
                self.load_students()
                self.load_statistics()
            else:
                QMessageBox.warning(self, "Erreur", "Impossible de supprimer l'étudiant")

    def export_students(self):
        """Exporte les étudiants vers Excel"""
        try:
            # Demander où sauvegarder le fichier
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Exporter les étudiants",
                f"etudiants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Fichiers Excel (*.xlsx)"
            )

            if not file_path:
                return

            # Récupérer les filtres actuels
            filters = {}
            if self.filter_site.currentText() != "Tous":
                filters['site_formation'] = self.filter_site.currentText()
            if self.filter_classe.currentText() != "Toutes":
                filters['classe'] = self.filter_classe.currentText()
            if self.filter_statut.currentText() != "Tous":
                filters['statut'] = self.filter_statut.currentText()

            # Récupérer les étudiants
            students = self.db_manager.get_etudiants(filters)

            if not students:
                QMessageBox.warning(self, "Aucun étudiant", "Aucun étudiant à exporter")
                return

            # Créer le fichier Excel
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Étudiants"

            # En-têtes
            headers = [
                "Matricule", "Nom", "Prénom", "Date de naissance", "Lieu de naissance",
                "Sexe", "Email", "Téléphone", "Classe", "Année académique",
                "Site", "Statut inscription", "Montant inscription", "Statut paiement"
            ]

            # Style des en-têtes
            header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Données
            for row, student in enumerate(students, 2):
                ws.cell(row=row, column=1, value=student.get('matricule', ''))
                ws.cell(row=row, column=2, value=student.get('nom', ''))
                ws.cell(row=row, column=3, value=student.get('prenom', ''))
                ws.cell(row=row, column=4, value=student.get('date_naissance', ''))
                ws.cell(row=row, column=5, value=student.get('lieu_naissance', ''))
                ws.cell(row=row, column=6, value=student.get('sexe', ''))
                ws.cell(row=row, column=7, value=student.get('email', ''))
                ws.cell(row=row, column=8, value=student.get('telephone', ''))
                ws.cell(row=row, column=9, value=student.get('classe', ''))
                ws.cell(row=row, column=10, value=student.get('annee_academique', ''))
                ws.cell(row=row, column=11, value=student.get('site_formation', ''))
                ws.cell(row=row, column=12, value=student.get('statut', ''))
                ws.cell(row=row, column=13, value=student.get('montant_inscription', ''))
                ws.cell(row=row, column=14, value=student.get('statut_paiement', ''))

            # Ajuster la largeur des colonnes
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

            # Sauvegarder
            wb.save(file_path)

            QMessageBox.information(
                self,
                "Succès",
                f"Export réussi!\n\n{len(students)} étudiant(s) exporté(s) vers:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export:\n{str(e)}")

    def load_classes_for_cert(self):
        """Charge les classes pour la génération de certificats"""
        site = self.cert_site.currentText()
        students = self.db_manager.get_etudiants({'site_formation': site, 'statut': 'Validé'})

        classes = sorted(set(s['classe'] for s in students if s['classe']))

        self.cert_classe.clear()
        self.cert_classe.addItems(classes)

    def preview_certificates(self):
        """Prévisualise les certificats"""
        QMessageBox.information(self, "Info", "Prévisualisation à implémenter")

    def generate_certificates(self):
        """Génère les certificats"""
        QMessageBox.information(self, "Info", "Génération de certificats à implémenter")

    def select_import_file(self):
        """Sélectionne un fichier Excel pour l'import"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Sélectionner un fichier Excel",
            "",
            "Fichiers Excel (*.xlsx)"
        )

        if file_path:
            self.import_file_label.setText(os.path.basename(file_path))
            self.import_file_path = file_path

    def download_template(self):
        """Télécharge le template Excel"""
        from import_liste import generer_template_excel

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Enregistrer le template",
            "template_import_etudiants.xlsx",
            "Fichiers Excel (*.xlsx)"
        )

        if file_path:
            generer_template_excel(file_path)
            QMessageBox.information(self, "Succès", f"Template enregistré: {file_path}")

    def import_excel(self):
        """Importe un fichier Excel"""
        if not hasattr(self, 'import_file_path'):
            QMessageBox.warning(self, "Erreur", "Veuillez sélectionner un fichier Excel")
            return

        QMessageBox.information(self, "Info", "Import Excel à implémenter complètement")

    def save_settings(self):
        """Enregistre les paramètres"""
        self.db_manager.set_config('annee_academique_actuelle', self.settings_annee.text())
        QMessageBox.information(self, "Succès", "Paramètres enregistrés avec succès")

    def backup_database(self):
        """Sauvegarde la base de données"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Sauvegarder la base de données",
            f"backup_enastic_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db",
            "Base de données (*.db)"
        )

        if file_path:
            if self.db_manager.backup_database(file_path):
                QMessageBox.information(self, "Succès", f"Sauvegarde créée: {file_path}")
            else:
                QMessageBox.warning(self, "Erreur", "Échec de la sauvegarde")

    def handle_logout(self):
        """Gère la déconnexion"""
        reply = QMessageBox.question(self, "Confirmation",
                                     "Voulez-vous vraiment vous déconnecter ?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            self.close()


class AddStudentDialog(QDialog):
    """Dialogue pour ajouter un nouvel étudiant"""

    def __init__(self, db_manager, user, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.user = user
        self.init_ui()

    def init_ui(self):
        """Initialise l'interface"""
        self.setWindowTitle("Ajouter un nouvel étudiant")
        self.setModal(True)
        self.setMinimumWidth(600)

        layout = QVBoxLayout()

        # Formulaire
        form = QFormLayout()

        # Matricule (optionnel - sera généré automatiquement si vide)
        self.matricule_input = QLineEdit()
        self.matricule_input.setPlaceholderText("Laissez vide pour génération automatique")
        form.addRow("Matricule (optionnel):", self.matricule_input)

        # Nom
        self.nom_input = QLineEdit()
        self.nom_input.setPlaceholderText("Nom de famille")
        form.addRow("Nom *:", self.nom_input)

        # Prénom
        self.prenom_input = QLineEdit()
        self.prenom_input.setPlaceholderText("Prénom(s)")
        form.addRow("Prénom *:", self.prenom_input)

        # Sexe
        self.sexe_combo = QComboBox()
        self.sexe_combo.addItems(["M", "F"])
        form.addRow("Sexe *:", self.sexe_combo)

        # Date de naissance
        self.date_naissance_input = QDateEdit()
        self.date_naissance_input.setCalendarPopup(True)
        self.date_naissance_input.setDate(QDate(2000, 1, 1))
        self.date_naissance_input.setDisplayFormat("dd/MM/yyyy")
        form.addRow("Date de naissance *:", self.date_naissance_input)

        # Lieu de naissance
        self.lieu_naissance_input = QLineEdit()
        self.lieu_naissance_input.setPlaceholderText("Ville, Pays")
        form.addRow("Lieu de naissance *:", self.lieu_naissance_input)

        # Email
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("exemple@email.com")
        form.addRow("Email:", self.email_input)

        # Téléphone
        self.telephone_input = QLineEdit()
        self.telephone_input.setPlaceholderText("+235 XX XX XX XX")
        form.addRow("Téléphone:", self.telephone_input)

        # Classe
        self.classe_combo = QComboBox()
        classes = ['MTIC1', 'MTIC2', 'MTIC3', 'TELECOMS1', 'TELECOMS2', 'TELECOMS3',
                   'ASRS1', 'ASRS2', 'ASRS3', 'GLRS1', 'GLRS2', 'GLRS3']
        self.classe_combo.addItems(classes)
        form.addRow("Classe *:", self.classe_combo)

        # Année académique
        self.annee_combo = QComboBox()
        current_year = datetime.now().year
        for year in range(current_year - 1, current_year + 3):
            self.annee_combo.addItem(f"{year}-{year+1}")
        self.annee_combo.setCurrentText(f"{current_year}-{current_year+1}")
        form.addRow("Année académique *:", self.annee_combo)

        # Site de formation
        self.site_combo = QComboBox()
        sites = ['Ndjamena', 'Sarh', 'Amdjarass']
        self.site_combo.addItems(sites)
        # Sélectionner le site de l'utilisateur par défaut
        if self.user.get('site') in sites:
            self.site_combo.setCurrentText(self.user['site'])
        form.addRow("Site de formation *:", self.site_combo)

        # Statut inscription
        self.statut_combo = QComboBox()
        self.statut_combo.addItems(['Validé', 'En attente', 'Refusé'])
        form.addRow("Statut inscription *:", self.statut_combo)

        # Montant inscription
        self.montant_input = QLineEdit()
        self.montant_input.setPlaceholderText("150000")
        self.montant_input.setText("150000")
        form.addRow("Montant inscription (FCFA) *:", self.montant_input)

        # Statut paiement
        self.paiement_combo = QComboBox()
        self.paiement_combo.addItems(['Payé', 'Partiel', 'Non payé'])
        self.paiement_combo.setCurrentText('Non payé')
        form.addRow("Statut paiement *:", self.paiement_combo)

        layout.addLayout(form)

        # Note
        note = QLabel("* Champs obligatoires")
        note.setStyleSheet("color: #6B7280; font-size: 11px; font-style: italic;")
        layout.addWidget(note)

        # Boutons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.save_student)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.setLayout(layout)

    def save_student(self):
        """Sauvegarde le nouvel étudiant"""
        try:
            # Validation
            if not self.nom_input.text().strip():
                QMessageBox.warning(self, "Erreur", "Le nom est obligatoire")
                return

            if not self.prenom_input.text().strip():
                QMessageBox.warning(self, "Erreur", "Le prénom est obligatoire")
                return

            if not self.lieu_naissance_input.text().strip():
                QMessageBox.warning(self, "Erreur", "Le lieu de naissance est obligatoire")
                return

            # Validation du montant
            try:
                montant = float(self.montant_input.text().strip())
                if montant < 0:
                    raise ValueError()
            except:
                QMessageBox.warning(self, "Erreur", "Le montant doit être un nombre valide")
                return

            # Préparer les données
            data = {
                'matricule': self.matricule_input.text().strip() or None,  # None pour auto-génération
                'nom': self.nom_input.text().strip().upper(),
                'prenom': self.prenom_input.text().strip().title(),
                'sexe': self.sexe_combo.currentText(),
                'date_naissance': self.date_naissance_input.date().toString("dd/MM/yyyy"),
                'lieu_naissance': self.lieu_naissance_input.text().strip(),
                'email': self.email_input.text().strip(),
                'telephone': self.telephone_input.text().strip(),
                'classe': self.classe_combo.currentText(),
                'annee_academique': self.annee_combo.currentText(),
                'site_formation': self.site_combo.currentText(),
                'statut': self.statut_combo.currentText(),
                'montant_inscription': montant,
                'statut_paiement': self.paiement_combo.currentText()
            }

            # Créer l'étudiant dans la base de données
            etudiant_id = self.db_manager.create_etudiant(data, self.user['id'])

            if etudiant_id:
                QMessageBox.information(
                    self,
                    "Succès",
                    f"Étudiant ajouté avec succès!\n\nMatricule: {data.get('matricule', 'Généré automatiquement')}"
                )
                self.accept()
            else:
                QMessageBox.critical(self, "Erreur", "Impossible de créer l'étudiant")

        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la création:\n{str(e)}")


class EditStudentDialog(AddStudentDialog):
    """Dialogue pour éditer un étudiant existant"""

    def __init__(self, db_manager, user, student, parent=None):
        self.student = student
        self.student_id = student['id']
        super().__init__(db_manager, user, parent)

    def init_ui(self):
        """Initialise l'interface avec les données de l'étudiant"""
        super().init_ui()

        # Changer le titre
        self.setWindowTitle("Modifier l'étudiant")

        # Pré-remplir les champs
        self.matricule_input.setText(self.student.get('matricule', ''))
        self.matricule_input.setReadOnly(True)  # Le matricule ne peut pas être modifié

        self.nom_input.setText(self.student.get('nom', ''))
        self.prenom_input.setText(self.student.get('prenom', ''))

        sexe = self.student.get('sexe', 'M')
        self.sexe_combo.setCurrentText(sexe)

        # Date de naissance
        date_str = self.student.get('date_naissance', '')
        if date_str:
            try:
                parts = date_str.split('/')
                if len(parts) == 3:
                    day, month, year = parts
                    self.date_naissance_input.setDate(QDate(int(year), int(month), int(day)))
            except:
                pass

        self.lieu_naissance_input.setText(self.student.get('lieu_naissance', ''))
        self.email_input.setText(self.student.get('email', ''))
        self.telephone_input.setText(self.student.get('telephone', ''))

        classe = self.student.get('classe', '')
        if classe:
            self.classe_combo.setCurrentText(classe)

        annee = self.student.get('annee_academique', '')
        if annee:
            self.annee_combo.setCurrentText(annee)

        site = self.student.get('site_formation', '')
        if site:
            self.site_combo.setCurrentText(site)

        statut = self.student.get('statut', 'En attente')
        self.statut_combo.setCurrentText(statut)

        montant = self.student.get('montant_inscription', '')
        if montant:
            self.montant_input.setText(str(montant))

        paiement = self.student.get('statut_paiement', 'Non payé')
        self.paiement_combo.setCurrentText(paiement)

    def save_student(self):
        """Met à jour l'étudiant existant"""
        try:
            # Validation
            if not self.nom_input.text().strip():
                QMessageBox.warning(self, "Erreur", "Le nom est obligatoire")
                return

            if not self.prenom_input.text().strip():
                QMessageBox.warning(self, "Erreur", "Le prénom est obligatoire")
                return

            if not self.lieu_naissance_input.text().strip():
                QMessageBox.warning(self, "Erreur", "Le lieu de naissance est obligatoire")
                return

            # Validation du montant
            try:
                montant = float(self.montant_input.text().strip())
                if montant < 0:
                    raise ValueError()
            except:
                QMessageBox.warning(self, "Erreur", "Le montant doit être un nombre valide")
                return

            # Préparer les données (sans le matricule qui ne peut pas être modifié)
            data = {
                'nom': self.nom_input.text().strip().upper(),
                'prenom': self.prenom_input.text().strip().title(),
                'sexe': self.sexe_combo.currentText(),
                'date_naissance': self.date_naissance_input.date().toString("dd/MM/yyyy"),
                'lieu_naissance': self.lieu_naissance_input.text().strip(),
                'email': self.email_input.text().strip(),
                'telephone': self.telephone_input.text().strip(),
                'classe': self.classe_combo.currentText(),
                'annee_academique': self.annee_combo.currentText(),
                'site_formation': self.site_combo.currentText(),
                'statut': self.statut_combo.currentText(),
                'montant_inscription': montant,
                'statut_paiement': self.paiement_combo.currentText()
            }

            # Mettre à jour l'étudiant
            if self.db_manager.update_etudiant(self.student_id, data, self.user['id']):
                QMessageBox.information(self, "Succès", "Étudiant modifié avec succès!")
                self.accept()
            else:
                QMessageBox.critical(self, "Erreur", "Impossible de modifier l'étudiant")

        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la modification:\n{str(e)}")
